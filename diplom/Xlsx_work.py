# Import pandas
import pandas as pd
import os
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

print(os.listdir())


# Assign spreadsheet filename to `file`
file = '123.xlsx'

# Load spreadsheet
xl = pd.ExcelFile(file)

# Print the sheet names
print(xl.sheet_names)

# Load a sheet into a DataFrame by name: df1
df1 = xl.parse('Лист1')
print(df1, "\n")
# Как индексация указана в таблице, так к элементами можно и обращаться
print(df1[1][0])

writer = pd.ExcelWriter('example.xlsx', engine='xlsxwriter')

DT = [[1,2,3],
      [4,5,6]]
DT1 = np.array([[1,2,3], [6,7,8]])
df1.to_excel(writer, 'Sheet1')
writer.save()

# Применение openpyxl
print(os.listdir())
wb = load_workbook('./Test_OPENPYXL.xlsx')
print(wb.sheetnames)
print(wb.active)

sheet = wb['Sheet1']

print(wb['Sheet1']['A1'].value)
c = sheet['B2']
print(c.row)
print(c.column)
print(c.coordinate)

print(sheet.cell(row=1, column=2).value, "\n")

for i in range(1,sheet.max_row+1):
      print(i, sheet.cell(row=i, column=2).value*10)
print("\n")

for i in range(1,sheet.max_column+1):
      print(i, sheet.cell(row=2, column=i).value*10)
print("\n")

# Получение буквы по индексу
print(get_column_letter(2))
print(column_index_from_string('A'), "\n")

# Фокусирование на области
for cellObj in sheet['A1':'C3']:
      for cell in cellObj:
            print(cell.coordinate, cell.value)
      print('---END---')
print("\n")

# Преобразование листа в
df = pd.DataFrame(sheet.values)
print(df)

# Put the sheet values in `data`
data = sheet.values

# Indicate the columns in the sheet values
cols = next(data)[1:]

# Convert your data to a list
data = list(data)

# Read in the data at index 0 for the indices
idx = [r[0] for r in data]

print(data)









